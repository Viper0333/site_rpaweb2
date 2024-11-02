from openpyxl import load_workbook                                                      # Importa a função para manipulação de arquivos Excel.
import os                                                                               # Importa o módulo para operações com o sistema operacional.
import datetime                                                                         # Importa o módulo para lidar com datas e horas.
import pandas as pd                                                                     # Importa o pandas, uma biblioteca para manipulação de dados.

converter_directory = os.path.dirname(os.path.abspath(__file__))                        # Obtém o diretório do arquivo .py atual

current_date_time = datetime.datetime.now()                                             # Captura a data e hora atuais
formatted_date = current_date_time.strftime("%d_%m_%Y")                                 # Formata a data como dia_mês_ano.

                                                                                        # Define o diretorio de arquivos convertidos com uma pasta de data atual
ABS_folder = os.path.join(converter_directory, 'data')                                  # Caminho da pasta onde estão os arquivos CSV.
converted_files_folder  = os.path.join(converter_directory, 'converted_files_folder ')  # Pasta para arquivos convertidos.
folder_name = f"{formatted_date}_Arquivos_Milimetro"                                    # Nome da subpasta com a data atual.
folder_path = os.path.join(converted_files_folder , folder_name)                        # Caminho completo para a nova subpasta.

                                                                                        # Cria a pasta para arquivos convertidos, se não existir
if not os.path.exists(converted_files_folder ):                                         # Verifica se a pasta de arquivos convertidos existe.
    os.makedirs(converted_files_folder )                                                # Cria a pasta se ela não existir.

                                                                                        # Cria a subpasta com a data atual, se não existir
if not os.path.exists(folder_path):                                                     # Verifica se a subpasta com data atual existe.
    os.makedirs(folder_path)                                                            # Cria a subpasta se ela não existir.

                                                                                        # Itera sobre cada arquivo na pasta "dados"
for amper_file in os.listdir(ABS_folder):                                               # Itera sobre cada arquivo na pasta "dados".
    file_path = os.path.join(ABS_folder, amper_file)                                    # Cria o caminho completo do arquivo.
    xlsx_file_name = os.path.splitext(amper_file)[0] + ".xlsx"                          # Define o nome para o arquivo Excel.
    file_path_xlsx = os.path.join(folder_path, xlsx_file_name)                          # Caminho completo do novo arquivo Excel.
    
                                                                                        # Verifica se o arquivo é CSV
    if amper_file.endswith(".csv"):                                                     # Confere se o arquivo tem extensão CSV.
        data = pd.read_csv(file_path, sep=";")                                          # Lê o arquivo CSV em um DataFrame.
        print(converted_files_folder )                                                  # Exibe o caminho dos arquivos convertidos.
                                                                                        
        data.to_excel(file_path_xlsx, index=False, engine='xlsxwriter')                 # Salva o DataFrame como arquivo XLSX na subpasta nome_pasta
        
    else:
        print(f"O arquivo {amper_file} não é um arquivo CSV. 
            Pulando para o próximo arquivo.")
        continue                                                                        # Pula arquivos que não são CSV.

    wb = load_workbook(file_path_xlsx)                                                  # Carrega o arquivo Excel recém-criado
    sheet1 = wb.active                                                                  # Seleciona a primeira aba do Excel.

                                                                                        # Adiciona cabeçalhos nas colunas F, G e H
    sheet1['F1'] = "Nivel_em_Amperes"                                                   # Coluna F: Cabeçalho para nível em amperes.
    sheet1['G1'] = "Nivel_em_Milimetros"                                                # Coluna G: Cabeçalho para nível em milímetros.
    sheet1['H1'] = "Nivel_em_Centimetros"                                               # Coluna H: Cabeçalho para nível em centímetros.

    bool_valido = [0]                                                                   # Inicializa uma variável para verificar valores válidos.
    counter = 0                                                                         # Contador para linhas processadas.

    for row in sheet1.iter_rows(min_col=6, max_col=6, values_only=False):               # Itera sobre cada célula na coluna F (Nível em Amperes)
        for cell in row:
            cell_value = cell.value                                                     # Valor da célula.
            cell_row = cell.row                                                         # Número da linha da célula.
            row_count = sheet1.max_row - 1                                              # Total de linhas na planilha, menos o cabeçalho.

            if cell_value == 0:
                counter += 1                                                            # Conta linhas com valor zero.
            
            if str(cell_value).isdigit() and int(cell_value) != 0:                      # Verifica se o valor é um número diferente de zero.
                
                result = int(cell_value) * 1450 - 340000                                # Calcula o nível em milímetros usando uma função linear
                millimeters = result / 1600

                                                                                        # Preenche as colunas G e H com os valores calculados
                sheet1.cell(row=cell_row, column=7, value=round(millimeters, 1))        # Coluna G: nível em milímetros.
                sheet1.cell(row=cell_row, column=8, value=round(millimeters / 10, 1))   # Coluna H: nível em centímetros.

                bool_valido = 1                                                         # Indica que foi encontrado um valor válido.
                counter += 1                                                            # Incrementa o contador para valores não nulos.
            
            if counter == row_count and bool_valido == 1:                               
                os.remove(file_path)                                                    # Remove o arquivo CSV após se a conversão for completa.
        
    wb.save(file_path_xlsx)                                                             # Salva as alterações no arquivo Excel.